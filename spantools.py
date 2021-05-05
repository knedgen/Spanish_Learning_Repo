#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#!/usr/bin/env python
# coding: utf-8

# In[37]:


import bs4
import requests
import time
from random import randint
from random import choice
from random import sample
from openpyxl import Workbook
from openpyxl import load_workbook
letters = ["a","b","c","d"]
result = requests.get('https://www.spanishdict.com/wordoftheday/1')
soup = bs4.BeautifulSoup(result.text,'lxml')


# In[41]:


class Spanish:
    
    def __init__(self,today,yesterday,week,month,year,span_comp,get_words,Quiz):
        
        self.today = today
        self.yesterday = yesterday
        self.week = week
        self.month = month
        self.year = year
        self.span_comp = span_comp
        self.get_words = get_words
        self.Quiz = Quiz
        
    def today(self):
        
        today = soup.select('h3')[0]
        totext = today.text
        today1 = soup.select('div.translation--3iXmZ8Jd')[0]
        totext1 = today1.text
        spex = soup.select("div.exampleSource--2I4LpW3B")[0]
        spextext = spex.text
        enex = soup.select('div.exampleTranslation--2w_JRz6o')[0]
        enextext = enex.text

        print("Today's word is:\n")
        print(f"{totext}")
        print(f"{totext1}\n")
        print("Example:\n")
        print(f"{spextext}")
        print(f"{enextext}")
   
    def yesterday(self):
    
        today = soup.select('h3')[1]
        totext = today.text
        today1 = soup.select('div.translation--3iXmZ8Jd')[1]
        totext1 = today1.text
        spex = soup.select("div.exampleSource--2I4LpW3B")[2]
        spextext = spex.text
        enex = soup.select('div.exampleTranslation--2w_JRz6o')[2]
        enextext = enex.text

        print("Yesterday's word was:\n")
        print(f"{totext}")
        print(f"{totext1}\n")
        print("Example:\n")
        print(f"{spextext}")
        print(f"{enextext}")
        
    def week(self):
    
        count = 0

        for i in range(0,7):

            result = requests.get('https://www.spanishdict.com/wordoftheday/1')
            soup = bs4.BeautifulSoup(result.text,'lxml')
            spword = soup.select('h3')[i]
            enword = soup.select('div.translation--3iXmZ8Jd')[i]
            count += 1


            print(f"{count}){spword.text}")
            print(f"{enword.text}\n")

    def month(self):

        count = 0

        for n in range(1,5):


            baseurl = ('https://www.spanishdict.com/wordoftheday/{}')
            scrapeurl = baseurl.format(n)

            result = requests.get(scrapeurl)
            soup = bs4.BeautifulSoup(result.text,'lxml')

            for i in range(0,10):

                spword = soup.select('h3')[i]
                enword = soup.select('div.translation--3iXmZ8Jd')[i]
                count += 1
                if count == 32:
                    break



                print(f"{count}){spword.text}")
                print(f"{enword.text}\n")
                
                
    def year(self):
    
        count = 0

        for n in range(1,38):


            baseurl = ('https://www.spanishdict.com/wordoftheday/{}')
            scrapeurl = baseurl.format(n)

            result = requests.get(scrapeurl)
            soup = bs4.BeautifulSoup(result.text,'lxml')

            for i in range(0,10):

                spword = soup.select('h3')[i]
                enword = soup.select('div.translation--3iXmZ8Jd')[i]
                count += 1

                if count == 366:
                    break


                print(f"{count}){spword.text}")
                print(f"{enword.text}\n")
                
    def span_comp(self):

        print("Welcome\n")
        print("Starting up...\n")
        time.sleep(1)


        while True:
                time.sleep(1)
                print("\nFunctions:")
                print("1. Today's Word")
                print("2. Yesterday's Word")
                print("3. Words the last week")
                print("4. Words the last month")
                print("5. Words the last year")
                print("6. Exit")
                print("\n")
                time.sleep(1)
                choice = int(input("Which function would you like to use? 1/2/3/4/5/6: "))
                print("\n")



                if choice == 1:
                    print("\nLoading...\n")
                    time.sleep(1)
                    Spanish.today(" ")
                    time.sleep(1)

                elif choice == 2:
                    print("\nLoading...\n")
                    time.sleep(1)
                    Spanish.yesterday(" ")
                    time.sleep(1)

                elif choice == 3:
                    print("\nLoading...\n")
                    time.sleep(1)
                    Spanish.week(" ")
                    time.sleep(1)

                elif choice == 4:
                    print("\nLoading...\n")
                    time.sleep(1)
                    Spanish.month(" ")
                    time.sleep(1)

                elif choice == 5:
                    print("\nLoading...\n")
                    time.sleep(1)
                    Spanish.year(" ")
                    time.sleep(1)

                elif choice == 6:
                    print("Shutting down...\n")
                    time.sleep(1)
                    print('Gracias!')
                    time.sleep(1)
                    break


                else:
                    print("Not a valid input")               

    def get_words(self):

        engwords = []
        spnwords = []


        for n in range(1,10):


            baseurl = ('https://www.spanishdict.com/wordoftheday/{}')
            scrapeurl = baseurl.format(n)

            result = requests.get(scrapeurl)
            soup = bs4.BeautifulSoup(result.text,'lxml')

            for i in range(0,10):

                spword = soup.select('h3')[i]
                spnwords.append(spword.text)

                enword = soup.select('div.translation--3iXmZ8Jd')[i]
                engwords.append(enword.text)


        spanengdict = {engwords[i]: spnwords[i] for i in range(len(engwords))}
        return spanengdict
    
    
    def Quiz(self):
    
        play = True
        ques = 1
        points = 0
        name = input("Please enter your name: ")
        print("\nGathering words, just a moment please....")


        quiz_words = Spanish.get_words(" ")

        while play is True:

            for x in range(1):
                choices = sample(list(quiz_words),4)
                rword = choice(choices)

                question = "What is the correct translation?"

                print(f"Total Points:{points}\n")
                print(f"Question {ques}) {question}\nWord: {rword}\n\n")
                print("Choices: ")

            options = list(quiz_words[x] for x in choices)

            for i in range(len(options)):

                op = str(letters[i]) + ")" + str(options[i])
                print(op)

            answer = input("\nEnter answer a,b,c or d:\nYour answer:")

            print("\n")

            if answer.lower() in letters:

                if quiz_words[rword] == options[letters.index(answer.lower())]:
                    points += 1

                    print("Correct, plus 1 point!")

                else:
                    print("Incorrect\n")
                    print(f"The correct answer is '{quiz_words[rword]}'") 

            else:
                print("Enter a valid input")


            print(f"Total Points:{points}\n")
            again = input("Play again? y/n: ")
            print("\n\n")

            if again == 'y' or again == 'Y' or again == 'yes' or again == 'Yes':
                ques += 1

                continue

            elif again == 'n' or again == 'N' or again == 'no' or again == 'No':
                
                #saves high scores to an excel file
                filename = 'hiscores.xlsx'
                new_row = [name,points,ques]

                #opens the workbook
                try:
                    wb = load_workbook('hiscores.xlsx')
                    #gets the first sheet
                    ws = wb.worksheets[0]
                #creates the file if it doesn't exist
                except:
                    headers_row = ['Name','Points','Questions']
                    wb = Workbook()
                    ws = wb.active
                    ws.append(headers_row)

                #add the info to the rows
                ws.append(new_row)
                wb.save("hiscores.xlsx")
                
                print(f"{name} scored {points} point(s) off of {ques} question(s)\n")
                print("Saving score...\n")
                time.sleep(1)
                print("Thanks for playing")
                play = False


# In[ ]:



